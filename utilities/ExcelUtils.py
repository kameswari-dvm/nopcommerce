import openpyxl


def open_excel_sheet(filename):
    workbook = openpyxl.load_workbook(filename)
    # sheet = workbook.get_sheet_by_name(sheet_name)
    sheet = workbook.sheetnames
    for i in range(0, sheet.__len__()):
        if sheet[i] == 'Sheet3':
            return workbook[sheet[i]]


def get_rows_count(filename):
    sheet = open_excel_sheet(filename)
    return sheet.max_row


def get_columns_count(filename):
    sheet = open_excel_sheet(filename)
    return sheet.max_column


def read_excel_file_data(filename, row_num, column_num):
    sheet = open_excel_sheet(filename)
    return sheet.cell(row=row_num, column=column_num).value


def write_data_into_excel_file(filename, row_num, column_num, data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook.sheetnames
    print(sheet.__len__())
    for i in range(0, sheet.__len__()):
        if sheet[i] == 'login_details':
            return sheet[i]
    sheet.cell(row=row_num, column=column_num).value = data
    workbook.save(filename)

# print(get_rows_count("C://pythonPracticeFiles//employee_data.xlsx", "Emp_details"))
